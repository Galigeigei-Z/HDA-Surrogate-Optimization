from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

try:
    from .session import HysysContext, ensure_context
except ImportError:
    from session import HysysContext, ensure_context

import numpy as np


EPS = 1e-9
STANDALONE_UTILITY_STREAM_NAMES = {
    "MP",
    "HP",
    "CW",
    "ChW",
    "Refrigeration",
    "LPSG",
}
THERMAL_OPERATION_KEYWORDS = (
    "heater",
    "cooler",
    "exchanger",
    "heatx",
    "heat exchanger",
    "reboiler",
    "condenser",
)


@dataclass(frozen=True)
class ThermalStream:
    name: str
    supply_temp_c: float
    target_temp_c: float
    heat_load_kw: float
    overall_u_kw_m2_k: float
    metadata: dict[str, str] = field(default_factory=dict)

    @property
    def stream_type(self) -> str:
        return "HOT" if self.supply_temp_c > self.target_temp_c else "COLD"

    @property
    def fcp_kw_per_k(self) -> float:
        delta_t = self.target_temp_c - self.supply_temp_c
        if math.isclose(delta_t, 0.0, abs_tol=EPS):
            raise ValueError(f"Stream {self.name} has zero temperature change.")
        return self.heat_load_kw / delta_t


@dataclass(frozen=True)
class ProblemRow:
    tleft_c: float
    tright_c: float
    temperature_difference_c: float
    active_hot_ids: tuple[int, ...]
    active_cold_ids: tuple[int, ...]
    sum_fcp_kw_per_k: float
    deficit_kw: float
    heat_in_kw: float
    heat_out_kw: float


@dataclass(frozen=True)
class CompositeCurvePoint:
    enthalpy_kw: float
    temperature_c: float
    active_stream_ids: tuple[int, ...]


@dataclass(frozen=True)
class AreaInterval:
    enthalpy_start_kw: float
    enthalpy_end_kw: float
    hot_in_c: float
    hot_out_c: float
    cold_in_c: float
    cold_out_c: float
    lmtd_c: float | None
    area_m2: float | None
    q_over_u_m2k: float
    hot_stream_ids: tuple[int, ...]
    cold_stream_ids: tuple[int, ...]

    @property
    def duty_kw(self) -> float:
        return abs(self.enthalpy_end_kw - self.enthalpy_start_kw)


@dataclass(frozen=True)
class SupertargetingResult:
    delta_tmin_c: float
    pinch_hot_c: float
    pinch_cold_c: float
    minimum_hot_utility_kw: float
    minimum_cold_utility_kw: float
    minimum_exchangers: int
    minimum_exchangers_above_pinch: int
    minimum_exchangers_below_pinch: int
    minimum_area_m2: float
    problem_table: tuple[ProblemRow, ...]
    area_intervals: tuple[AreaInterval, ...]


@dataclass(frozen=True)
class NotebookAreaSegment:
    segment: int
    hot_in_c: float
    cold_in_c: float
    hot_out_c: float
    cold_out_c: float
    delta_ta_c: float
    delta_tb_c: float
    lmtd_c: float
    hot_q_over_u_m2k: float
    cold_q_over_u_m2k: float
    area_m2: float
    hot_streams: tuple[tuple[int, float, float], ...]
    cold_streams: tuple[tuple[int, float, float], ...]


@dataclass(frozen=True)
class NotebookReplayResult:
    delta_tmin_c: float
    pinch_hot_c: float
    minimum_hot_utility_kw: float
    minimum_cold_utility_kw: float
    total_area_m2: float
    segments: tuple[NotebookAreaSegment, ...]


class HysysValueSource:
    def read(self, *, context: HysysContext) -> float:
        raise NotImplementedError


@dataclass(frozen=True)
class ConstantValue(HysysValueSource):
    value: float

    def read(self, *, context: HysysContext) -> float:
        return self.value


@dataclass(frozen=True)
class MaterialTemperature(HysysValueSource):
    stream_name: str

    def read(self, *, context: HysysContext) -> float:
        return float(context.flowsheet.MaterialStreams.Item(self.stream_name).Temperature.Value)


@dataclass(frozen=True)
class EnergyDuty(HysysValueSource):
    energy_stream_name: str
    absolute: bool = True
    scale: float = 1.0

    def read(self, *, context: HysysContext) -> float:
        value = float(context.flowsheet.EnergyStreams.Item(self.energy_stream_name).HeatFlow.Value)
        if self.absolute:
            value = abs(value)
        return value * self.scale


@dataclass(frozen=True)
class MaterialEnthalpyDelta(HysysValueSource):
    inlet_stream_name: str
    outlet_stream_name: str
    absolute: bool = True
    scale: float = 1.0

    def read(self, *, context: HysysContext) -> float:
        inlet = context.flowsheet.MaterialStreams.Item(self.inlet_stream_name)
        outlet = context.flowsheet.MaterialStreams.Item(self.outlet_stream_name)
        inlet_kw = float(inlet.MolarFlow.Value) * float(inlet.MolarEnthalpy.Value)
        outlet_kw = float(outlet.MolarFlow.Value) * float(outlet.MolarEnthalpy.Value)
        value = outlet_kw - inlet_kw
        if self.absolute:
            value = abs(value)
        return value * self.scale


@dataclass(frozen=True)
class HysysStreamSpec:
    name: str
    supply_temperature: HysysValueSource
    target_temperature: HysysValueSource
    heat_load_kw: HysysValueSource
    overall_u_kw_m2_k: float
    metadata: dict[str, str] = field(default_factory=dict)

    def read(self, *, context: HysysContext) -> ThermalStream:
        return ThermalStream(
            name=self.name,
            supply_temp_c=self.supply_temperature.read(context=context),
            target_temp_c=self.target_temperature.read(context=context),
            heat_load_kw=self.heat_load_kw.read(context=context),
            overall_u_kw_m2_k=self.overall_u_kw_m2_k,
            metadata=dict(self.metadata),
        )


def _safe_getattr(obj: Any, name: str) -> Any | None:
    try:
        return getattr(obj, name)
    except Exception:
        return None


def _extract_value(obj: Any) -> Any:
    if obj is None:
        return None
    try:
        return obj.Value
    except Exception:
        return obj


def _coerce_float(value: Any) -> float | None:
    raw = _extract_value(value)
    if raw is None:
        return None
    try:
        return float(raw)
    except Exception:
        return None


def _iter_collection_items(collection: Any) -> list[Any]:
    if collection is None:
        return []
    names = _safe_getattr(collection, "Names")
    if names is not None:
        try:
            return [collection.Item(name) for name in names]
        except Exception:
            pass
    try:
        return list(collection)
    except Exception:
        return []


def _object_name(obj: Any) -> str | None:
    name = _safe_getattr(obj, "Name")
    if name is None:
        return None
    try:
        return str(name)
    except Exception:
        return None


def _read_stream_temp_c(stream: Any) -> float | None:
    temperature = _safe_getattr(stream, "Temperature")
    return _coerce_float(temperature)


def _read_stream_cp_candidates(stream: Any) -> dict[str, float]:
    candidates = {
        "HeatCapacityFlow": _coerce_float(_safe_getattr(stream, "HeatCapacityFlow")),
        "Cp": _coerce_float(_safe_getattr(stream, "Cp")),
        "CpMass": _coerce_float(_safe_getattr(stream, "CpMass")),
        "CpMolar": _coerce_float(_safe_getattr(stream, "CpMolar")),
        "MassHeatCapacity": _coerce_float(_safe_getattr(stream, "MassHeatCapacity")),
        "MolarHeatCapacity": _coerce_float(_safe_getattr(stream, "MolarHeatCapacity")),
    }
    return {key: value for key, value in candidates.items() if value is not None}


def _collect_named_streams(candidate: Any, *, expected_kind: str | None = None) -> list[Any]:
    streams: list[Any] = []
    if candidate is None:
        return streams

    name = _object_name(candidate)
    type_name = str(_safe_getattr(candidate, "TypeName") or "").lower()
    if name and (not expected_kind or expected_kind in type_name):
        streams.append(candidate)
        return streams

    for item in _iter_collection_items(candidate):
        item_name = _object_name(item)
        item_type = str(_safe_getattr(item, "TypeName") or "").lower()
        if item_name and (not expected_kind or expected_kind in item_type):
            streams.append(item)
    return streams


def _dedupe_by_name(objects: Sequence[Any]) -> list[Any]:
    deduped: list[Any] = []
    seen: set[str] = set()
    for obj in objects:
        name = _object_name(obj)
        if not name or name in seen:
            continue
        seen.add(name)
        deduped.append(obj)
    return deduped


def _operation_material_streams(operation: Any) -> tuple[list[Any], list[Any]]:
    feed_attrs = [
        "Feed",
        "FeedStream",
        "FeedStreams",
        "Inlet",
        "InletStream",
        "InletStreams",
        "MaterialStreams",
        "TubeInlet",
        "ShellInlet",
        "ProcessInlet",
    ]
    product_attrs = [
        "Product",
        "ProductStream",
        "ProductStreams",
        "Outlet",
        "OutletStream",
        "OutletStreams",
        "TubeOutlet",
        "ShellOutlet",
        "ProcessOutlet",
    ]

    feeds: list[Any] = []
    products: list[Any] = []
    for attr in feed_attrs:
        feeds.extend(_collect_named_streams(_safe_getattr(operation, attr), expected_kind="material"))
    for attr in product_attrs:
        products.extend(_collect_named_streams(_safe_getattr(operation, attr), expected_kind="material"))

    if not feeds or not products:
        ports = _safe_getattr(operation, "Ports")
        for port in _iter_collection_items(ports):
            port_name = str(_object_name(port) or "").lower()
            connected = (
                _safe_getattr(port, "ConnectedObject")
                or _safe_getattr(port, "AttachedObject")
                or _safe_getattr(port, "Stream")
            )
            material_streams = _collect_named_streams(connected, expected_kind="material")
            if not material_streams:
                continue
            if any(keyword in port_name for keyword in ("feed", "inlet", "in")):
                feeds.extend(material_streams)
            elif any(keyword in port_name for keyword in ("product", "outlet", "out")):
                products.extend(material_streams)

    return _dedupe_by_name(feeds), _dedupe_by_name(products)


def _operation_energy_streams(operation: Any, *, context: HysysContext) -> list[Any]:
    energy_attrs = [
        "EnergyStream",
        "EnergyStreams",
        "Duty",
        "DutyStream",
        "DutyStreams",
        "QStream",
        "QStreams",
    ]
    streams: list[Any] = []
    for attr in energy_attrs:
        streams.extend(_collect_named_streams(_safe_getattr(operation, attr), expected_kind="energy"))

    if not streams:
        op_name = str(_object_name(operation) or "")
        prefix = f"{op_name}-"
        for energy_name in getattr(context.flowsheet.EnergyStreams, "Names", []):
            if str(energy_name).startswith(prefix):
                try:
                    streams.append(context.flowsheet.EnergyStreams.Item(energy_name))
                except Exception:
                    continue
    return _dedupe_by_name(streams)


def _operation_u_value(operation: Any) -> float | None:
    u_attrs = [
        "OverallHeatTransferCoefficient",
        "OverallHeatTransferCoeff",
        "UA",
        "U",
    ]
    for attr in u_attrs:
        value = _coerce_float(_safe_getattr(operation, attr))
        if value is not None and value > 0.0:
            return value
    return None


def _looks_like_thermal_operation(operation: Any) -> bool:
    text = f"{_object_name(operation) or ''} {_safe_getattr(operation, 'TypeName') or ''}".lower()
    return any(keyword in text for keyword in THERMAL_OPERATION_KEYWORDS)


# Verify the uncertain stream names against the actual HYSYS case before production use.
DEFAULT_CFP04_HYSYS_SPECS: tuple[HysysStreamSpec, ...] = (
    HysysStreamSpec(
        name="7_To_28 (E-100)",
        supply_temperature=MaterialTemperature("S7"),
        target_temperature=MaterialTemperature("S28"),
        heat_load_kw=MaterialEnthalpyDelta("S7", "S28"),
        overall_u_kw_m2_k=0.15,
        metadata={"kind": "process"},
    ),
    HysysStreamSpec(
        name="7_To_28 (E-102)",
        supply_temperature=MaterialTemperature("S28"),
        target_temperature=MaterialTemperature("S9"),
        heat_load_kw=MaterialEnthalpyDelta("S28", "S9"),
        overall_u_kw_m2_k=0.60,
        metadata={"kind": "process", "note": "Verify S28->S9 mapping."},
    ),
    HysysStreamSpec(
        name="1Biphenyl_To_Bipheny-Final",
        supply_temperature=MaterialTemperature("S34"),
        target_temperature=MaterialTemperature("Bipheny-Final"),
        heat_load_kw=EnergyDuty("Cooler4-CW"),
        overall_u_kw_m2_k=0.60,
        metadata={"kind": "process"},
    ),
    HysysStreamSpec(
        name="1Benzene_To_Benzene Product",
        supply_temperature=MaterialTemperature("S27"),
        target_temperature=MaterialTemperature("Benzene Product"),
        heat_load_kw=EnergyDuty("Cooler3-CW"),
        overall_u_kw_m2_k=0.60,
        metadata={"kind": "process"},
    ),
    HysysStreamSpec(
        name="To Condenser@COL1",
        supply_temperature=MaterialTemperature("S23"),
        target_temperature=ConstantValue(-23.7),
        heat_load_kw=EnergyDuty("T1-Ref"),
        overall_u_kw_m2_k=0.60,
        metadata={"kind": "utility_hot", "note": "Target temperature is workbook-based."},
    ),
    HysysStreamSpec(
        name="To Condenser@COL2",
        supply_temperature=MaterialTemperature("S26"),
        target_temperature=ConstantValue(108.5),
        heat_load_kw=EnergyDuty("T2-CW"),
        overall_u_kw_m2_k=0.60,
        metadata={"kind": "utility_hot"},
    ),
    HysysStreamSpec(
        name="To Condenser@COL3",
        supply_temperature=MaterialTemperature("S30"),
        target_temperature=ConstantValue(138.4),
        heat_load_kw=EnergyDuty("T3-CW"),
        overall_u_kw_m2_k=0.60,
        metadata={"kind": "utility_hot"},
    ),
    HysysStreamSpec(
        name="4_To_6 (E-101)",
        supply_temperature=MaterialTemperature("S4"),
        target_temperature=MaterialTemperature("S6"),
        heat_load_kw=MaterialEnthalpyDelta("S4", "S6"),
        overall_u_kw_m2_k=0.15,
        metadata={"kind": "process", "note": "Verify S4->S6 mapping."},
    ),
    HysysStreamSpec(
        name="To Reboiler@COL1",
        supply_temperature=ConstantValue(187.5),
        target_temperature=ConstantValue(189.6),
        heat_load_kw=EnergyDuty("T1-HP"),
        overall_u_kw_m2_k=0.50,
        metadata={"kind": "utility_cold", "note": "Temperatures are workbook-based."},
    ),
    HysysStreamSpec(
        name="To Reboiler@COL2",
        supply_temperature=ConstantValue(146.0),
        target_temperature=ConstantValue(146.7),
        heat_load_kw=EnergyDuty("T2-MP"),
        overall_u_kw_m2_k=0.50,
        metadata={"kind": "utility_cold"},
    ),
    HysysStreamSpec(
        name="To Reboiler@COL3",
        supply_temperature=ConstantValue(295.3),
        target_temperature=ConstantValue(295.9),
        heat_load_kw=EnergyDuty("T3-FH"),
        overall_u_kw_m2_k=0.50,
        metadata={"kind": "utility_cold"},
    ),
    HysysStreamSpec(
        name="MP",
        supply_temperature=ConstantValue(175.0),
        target_temperature=ConstantValue(174.9),
        heat_load_kw=EnergyDuty("T2-MP"),
        overall_u_kw_m2_k=2.50,
        metadata={"kind": "utility_hot"},
    ),
    HysysStreamSpec(
        name="HP",
        supply_temperature=ConstantValue(250.0),
        target_temperature=ConstantValue(249.9),
        heat_load_kw=EnergyDuty("Heater1-HP"),
        overall_u_kw_m2_k=2.20,
        metadata={"kind": "utility_hot"},
    ),
    HysysStreamSpec(
        name="CW",
        supply_temperature=ConstantValue(30.0),
        target_temperature=ConstantValue(35.0),
        heat_load_kw=EnergyDuty("Cooler1-CW"),
        overall_u_kw_m2_k=0.50,
        metadata={"kind": "utility_cold"},
    ),
    HysysStreamSpec(
        name="Refrigeration",
        supply_temperature=ConstantValue(-40.0),
        target_temperature=ConstantValue(-35.0),
        heat_load_kw=EnergyDuty("T1-Ref"),
        overall_u_kw_m2_k=2.00,
        metadata={"kind": "utility_cold"},
    ),
)


def _float(value: object, *, field_name: str) -> float:
    if value is None or value == "":
        raise ValueError(f"Missing value for {field_name}.")
    return float(value)


def _clean_header(header: object) -> str:
    return str(header).strip() if header is not None else ""


def _pick_column(headers: Sequence[str], *candidates: str) -> int:
    normalized = {header.lower(): idx for idx, header in enumerate(headers)}
    for candidate in candidates:
        idx = normalized.get(candidate.lower())
        if idx is not None:
            return idx
    raise KeyError(f"Expected one of columns {candidates}, got {headers}.")


def _rows_to_streams(rows: Sequence[Sequence[object]]) -> list[ThermalStream]:
    if not rows:
        raise ValueError("Input table is empty.")

    headers = [_clean_header(header) for header in rows[0]]
    idx_name = _pick_column(headers, "Stream Information", "Stream", "Name")
    idx_ts = _pick_column(headers, "Supply Temperture (°C)", "Supply Temperature (°C)", "Ts")
    idx_tt = _pick_column(headers, "Target Temperature (°C)", "Tt")
    idx_q = _pick_column(headers, "Heat Load (kW)", "Q (kW)", "Heat Load")
    idx_u = _pick_column(headers, "U (KW/m2.K)", "U (kW/m2.K)", "U")

    streams: list[ThermalStream] = []
    for raw_row in rows[1:]:
        if not raw_row or all(value in (None, "") for value in raw_row):
            continue
        name = str(raw_row[idx_name]).strip()
        if not name:
            continue
        streams.append(
            ThermalStream(
                name=name,
                supply_temp_c=_float(raw_row[idx_ts], field_name=f"{name}.Ts"),
                target_temp_c=_float(raw_row[idx_tt], field_name=f"{name}.Tt"),
                heat_load_kw=abs(_float(raw_row[idx_q], field_name=f"{name}.Q")),
                overall_u_kw_m2_k=_float(raw_row[idx_u], field_name=f"{name}.U"),
            )
        )
    if not streams:
        raise ValueError("No valid thermal streams were loaded.")
    return streams


def read_streams_from_csv(path: str | Path) -> list[ThermalStream]:
    with Path(path).open("r", newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    return _rows_to_streams(rows)


def read_streams_from_records(records: Sequence[Mapping[str, object]]) -> list[ThermalStream]:
    if not records:
        raise ValueError("Input records are empty.")
    headers = list(records[0].keys())
    rows: list[list[object]] = [headers]
    for record in records:
        rows.append([record.get(header) for header in headers])
    return _rows_to_streams(rows)


def read_streams_from_xlsx(path: str | Path, sheet_name: str | None = None) -> list[ThermalStream]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx input files.") from exc

    workbook = load_workbook(filename=Path(path), data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    return _rows_to_streams(rows)


def read_streams_from_hysys(
    specs: Sequence[HysysStreamSpec],
    *,
    context: HysysContext | None = None,
) -> list[ThermalStream]:
    ctx = ensure_context(context)
    return [spec.read(context=ctx) for spec in specs]


def read_default_cfp04_hysys_streams(
    *,
    context: HysysContext | None = None,
) -> list[ThermalStream]:
    return read_streams_from_hysys(DEFAULT_CFP04_HYSYS_SPECS, context=context)


def scan_hysys_flowsheet_for_thermal_streams(
    *,
    context: HysysContext | None = None,
    fallback_u_kw_m2_k: float = 0.6,
) -> list[ThermalStream]:
    ctx = ensure_context(context)
    thermal_streams: list[ThermalStream] = []

    operation_names = getattr(ctx.flowsheet.Operations, "Names", [])
    for operation_name in operation_names:
        try:
            operation = ctx.flowsheet.Operations.Item(operation_name)
        except Exception:
            continue
        if not _looks_like_thermal_operation(operation):
            continue

        feed_streams, product_streams = _operation_material_streams(operation)
        energy_streams = _operation_energy_streams(operation, context=ctx)
        if not feed_streams or not product_streams:
            continue

        inlet_stream = feed_streams[0]
        outlet_stream = product_streams[0]
        ts = _read_stream_temp_c(inlet_stream)
        tt = _read_stream_temp_c(outlet_stream)
        if ts is None or tt is None or math.isclose(ts, tt, abs_tol=EPS):
            continue

        duty_kw = None
        energy_name = None
        for energy_stream in energy_streams:
            energy_name = _object_name(energy_stream)
            duty_kw = _coerce_float(_safe_getattr(energy_stream, "HeatFlow"))
            if duty_kw is not None:
                duty_kw = abs(duty_kw)
                break
        if duty_kw is None:
            duty_kw = abs(float(_safe_getattr(operation, "DutyValue") or 0.0)) or None
        if duty_kw is None or duty_kw <= 0.0:
            continue

        cp_metadata = _read_stream_cp_candidates(inlet_stream)
        overall_u = _operation_u_value(operation) or fallback_u_kw_m2_k
        metadata = {
            "kind": "auto_scanned",
            "operation_name": str(operation_name),
            "operation_type": str(_safe_getattr(operation, "TypeName") or ""),
            "inlet_stream_name": str(_object_name(inlet_stream) or ""),
            "outlet_stream_name": str(_object_name(outlet_stream) or ""),
            "energy_stream_name": str(energy_name or ""),
            "cp_source": ",".join(sorted(cp_metadata.keys())),
        }
        for key, value in cp_metadata.items():
            metadata[f"cp_{key}"] = f"{value}"
        thermal_streams.append(
            ThermalStream(
                name=str(operation_name),
                supply_temp_c=float(ts),
                target_temp_c=float(tt),
                heat_load_kw=float(duty_kw),
                overall_u_kw_m2_k=float(overall_u),
                metadata=metadata,
            )
        )
    return thermal_streams


def filter_out_standalone_utility_streams(streams: Sequence[ThermalStream]) -> list[ThermalStream]:
    return [stream for stream in streams if stream.name not in STANDALONE_UTILITY_STREAM_NAMES]


def stream_records(streams: Sequence[ThermalStream]) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for index, stream in enumerate(streams, start=1):
        records.append(
            {
                "stream_index": index,
                "name": stream.name,
                "stream_type": stream.stream_type,
                "supply_temp_c": stream.supply_temp_c,
                "target_temp_c": stream.target_temp_c,
                "heat_load_kw": stream.heat_load_kw,
                "overall_u_kw_m2_k": stream.overall_u_kw_m2_k,
                "fcp_kw_per_k": stream.fcp_kw_per_k,
                **stream.metadata,
            }
        )
    return records


def spec_records(specs: Sequence[HysysStreamSpec]) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for index, spec in enumerate(specs, start=1):
        records.append(
            {
                "spec_index": index,
                "name": spec.name,
                "u_kw_m2_k": spec.overall_u_kw_m2_k,
                **spec.metadata,
            }
        )
    return records


def _build_notebook_table(streams: Sequence[ThermalStream]) -> list[dict[str, object]]:
    table: list[dict[str, object]] = []
    for stream_id, stream in enumerate(streams, start=1):
        table.append(
            {
                "id": stream_id,
                "name": stream.name,
                "Ts": float(stream.supply_temp_c),
                "Tt": float(stream.target_temp_c),
                "Q": float(stream.heat_load_kw),
                "U": float(stream.overall_u_kw_m2_k),
                "FCp": round(stream.heat_load_kw / (stream.target_temp_c - stream.supply_temp_c), 2),
                "Stream Type": stream.stream_type,
            }
        )
    return table


def _unique_sorted_desc(values: Iterable[float]) -> list[float]:
    unique = sorted({round(value, 10) for value in values}, reverse=True)
    return [float(value) for value in unique]


def _active_stream_ids(
    streams: Sequence[ThermalStream],
    tleft_c: float,
    tright_c: float,
) -> tuple[tuple[int, ...], tuple[int, ...]]:
    hot_ids: list[int] = []
    cold_ids: list[int] = []
    for idx, stream in enumerate(streams):
        if stream.stream_type == "COLD":
            if stream.supply_temp_c <= tleft_c < stream.target_temp_c:
                cold_ids.append(idx)
        else:
            if stream.target_temp_c <= tright_c < stream.supply_temp_c:
                hot_ids.append(idx)
    return tuple(hot_ids), tuple(cold_ids)


def build_problem_table(streams: Sequence[ThermalStream], delta_tmin_c: float) -> tuple[ProblemRow, ...]:
    hot_streams = [stream for stream in streams if stream.stream_type == "HOT"]
    cold_streams = [stream for stream in streams if stream.stream_type == "COLD"]

    temp_left = _unique_sorted_desc(
        [stream.supply_temp_c for stream in cold_streams]
        + [stream.target_temp_c for stream in cold_streams]
        + [stream.supply_temp_c - delta_tmin_c for stream in hot_streams]
        + [stream.target_temp_c - delta_tmin_c for stream in hot_streams]
    )
    temp_right = _unique_sorted_desc(
        [stream.supply_temp_c for stream in hot_streams]
        + [stream.target_temp_c for stream in hot_streams]
        + [stream.supply_temp_c + delta_tmin_c for stream in cold_streams]
        + [stream.target_temp_c + delta_tmin_c for stream in cold_streams]
    )
    if len(temp_left) != len(temp_right):
        raise RuntimeError(
            f"Temperature ladders are inconsistent: left={len(temp_left)} right={len(temp_right)}."
        )

    provisional_rows: list[dict[str, object]] = []
    for idx, (tleft_c, tright_c) in enumerate(zip(temp_left, temp_right)):
        difference_c = 0.0 if idx == 0 else temp_left[idx - 1] - tleft_c
        hot_ids, cold_ids = _active_stream_ids(streams, tleft_c, tright_c)
        sum_fcp = sum(streams[stream_id].fcp_kw_per_k for stream_id in (*hot_ids, *cold_ids))
        deficit_kw = sum_fcp * difference_c
        provisional_rows.append(
            {
                "tleft_c": tleft_c,
                "tright_c": tright_c,
                "temperature_difference_c": difference_c,
                "active_hot_ids": hot_ids,
                "active_cold_ids": cold_ids,
                "sum_fcp_kw_per_k": sum_fcp,
                "deficit_kw": deficit_kw,
            }
        )

    accumulated_output: list[float] = []
    running_output = 0.0
    for row in provisional_rows:
        running_output += -float(row["deficit_kw"])
        accumulated_output.append(running_output)

    offset = abs(min(accumulated_output, default=0.0))
    problem_rows: list[ProblemRow] = []
    for idx, row in enumerate(provisional_rows):
        heat_in_kw = offset + (0.0 if idx == 0 else accumulated_output[idx - 1])
        heat_out_kw = offset + accumulated_output[idx]
        problem_rows.append(
            ProblemRow(
                tleft_c=float(row["tleft_c"]),
                tright_c=float(row["tright_c"]),
                temperature_difference_c=float(row["temperature_difference_c"]),
                active_hot_ids=tuple(row["active_hot_ids"]),  # type: ignore[arg-type]
                active_cold_ids=tuple(row["active_cold_ids"]),  # type: ignore[arg-type]
                sum_fcp_kw_per_k=float(row["sum_fcp_kw_per_k"]),
                deficit_kw=float(row["deficit_kw"]),
                heat_in_kw=heat_in_kw,
                heat_out_kw=heat_out_kw,
            )
        )
    return tuple(problem_rows)


def _find_pinch(problem_rows: Sequence[ProblemRow], delta_tmin_c: float) -> tuple[float, float]:
    pinch_row = min(problem_rows, key=lambda row: abs(row.heat_out_kw))
    return pinch_row.tright_c, pinch_row.tright_c - delta_tmin_c


def _build_curve(
    streams: Sequence[ThermalStream],
    problem_rows: Sequence[ProblemRow],
    *,
    curve_type: str,
    minimum_cold_utility_kw: float,
) -> tuple[CompositeCurvePoint, ...]:
    if curve_type not in {"HOT", "COLD"}:
        raise ValueError(f"Unsupported curve type {curve_type}.")

    if curve_type == "HOT":
        valid_temperatures = {
            round(stream.supply_temp_c, 10)
            for stream in streams
            if stream.stream_type == "HOT"
        } | {
            round(stream.target_temp_c, 10)
            for stream in streams
            if stream.stream_type == "HOT"
        }
    else:
        valid_temperatures = {
            round(stream.supply_temp_c, 10)
            for stream in streams
            if stream.stream_type == "COLD"
        } | {
            round(stream.target_temp_c, 10)
            for stream in streams
            if stream.stream_type == "COLD"
        }

    points: list[CompositeCurvePoint] = []
    cumulative_dh = 0.0
    reversed_rows = list(reversed(problem_rows))
    for row in reversed_rows:
        active_ids = row.active_hot_ids if curve_type == "HOT" else row.active_cold_ids
        sum_fcp = sum(streams[stream_id].fcp_kw_per_k for stream_id in active_ids)
        enthalpy_kw = abs(cumulative_dh)
        if curve_type == "COLD":
            enthalpy_kw += minimum_cold_utility_kw
        temperature_c = row.tright_c if curve_type == "HOT" else row.tleft_c
        if round(temperature_c, 10) in valid_temperatures:
            points.append(
                CompositeCurvePoint(
                    enthalpy_kw=enthalpy_kw,
                    temperature_c=temperature_c,
                    active_stream_ids=tuple(active_ids),
                )
            )
        cumulative_dh += row.temperature_difference_c * sum_fcp
    return tuple(points)


def _prepare_curve(curve: Sequence[CompositeCurvePoint]) -> tuple[list[float], list[float]]:
    xs: list[float] = []
    ys: list[float] = []
    for point in curve:
        x = point.enthalpy_kw
        y = point.temperature_c
        if xs and math.isclose(x, xs[-1], abs_tol=EPS):
            ys[-1] = y
        else:
            xs.append(x)
            ys.append(y)
    return xs, ys


def _interpolate_temperature(enthalpy_kw: float, curve: Sequence[CompositeCurvePoint]) -> float:
    xs, ys = _prepare_curve(curve)
    if not xs:
        raise ValueError("Composite curve is empty.")
    if enthalpy_kw <= xs[0]:
        return ys[0]
    if enthalpy_kw >= xs[-1]:
        return ys[-1]
    for idx in range(len(xs) - 1):
        x1 = xs[idx]
        x2 = xs[idx + 1]
        if x1 - EPS <= enthalpy_kw <= x2 + EPS:
            if math.isclose(x1, x2, abs_tol=EPS):
                return ys[idx + 1]
            fraction = (enthalpy_kw - x1) / (x2 - x1)
            return ys[idx] + fraction * (ys[idx + 1] - ys[idx])
    return ys[-1]


def _overlap_length(a1: float, a2: float, b1: float, b2: float) -> float:
    low = max(min(a1, a2), min(b1, b2))
    high = min(max(a1, a2), max(b1, b2))
    return max(0.0, high - low)


def _active_curve_candidates(
    curve: Sequence[CompositeCurvePoint],
    enthalpy_start_kw: float,
    enthalpy_end_kw: float,
) -> tuple[int, ...]:
    candidates: set[int] = set()
    for idx in range(len(curve) - 1):
        segment_start = curve[idx].enthalpy_kw
        segment_end = curve[idx + 1].enthalpy_kw
        if segment_start < enthalpy_end_kw - EPS and segment_end > enthalpy_start_kw + EPS:
            candidates.update(curve[idx].active_stream_ids)
    return tuple(sorted(candidates))


def build_area_intervals(
    streams: Sequence[ThermalStream],
    problem_rows: Sequence[ProblemRow],
    minimum_cold_utility_kw: float,
) -> tuple[AreaInterval, ...]:
    hot_curve = _build_curve(
        streams,
        problem_rows,
        curve_type="HOT",
        minimum_cold_utility_kw=minimum_cold_utility_kw,
    )
    cold_curve = _build_curve(
        streams,
        problem_rows,
        curve_type="COLD",
        minimum_cold_utility_kw=minimum_cold_utility_kw,
    )

    enthalpy_points = sorted(
        {
            round(point.enthalpy_kw, 10)
            for point in hot_curve
        }
        | {
            round(point.enthalpy_kw, 10)
            for point in cold_curve
        }
    )
    intervals: list[AreaInterval] = []
    for idx in range(len(enthalpy_points) - 1):
        h1 = float(enthalpy_points[idx])
        h2 = float(enthalpy_points[idx + 1])
        if math.isclose(h1, h2, abs_tol=EPS):
            continue

        hot_t1 = _interpolate_temperature(h1, hot_curve)
        hot_t2 = _interpolate_temperature(h2, hot_curve)
        cold_t1 = _interpolate_temperature(h1, cold_curve)
        cold_t2 = _interpolate_temperature(h2, cold_curve)

        hot_in_c = max(hot_t1, hot_t2)
        hot_out_c = min(hot_t1, hot_t2)
        cold_in_c = min(cold_t1, cold_t2)
        cold_out_c = max(cold_t1, cold_t2)

        hot_candidates = _active_curve_candidates(hot_curve, h1, h2)
        cold_candidates = _active_curve_candidates(cold_curve, h1, h2)
        hot_active = tuple(
            stream_id
            for stream_id in hot_candidates
            if streams[stream_id].stream_type == "HOT"
            and _overlap_length(
                streams[stream_id].supply_temp_c,
                streams[stream_id].target_temp_c,
                hot_in_c,
                hot_out_c,
            )
            > EPS
        )
        cold_active = tuple(
            stream_id
            for stream_id in cold_candidates
            if streams[stream_id].stream_type == "COLD"
            and _overlap_length(
                streams[stream_id].supply_temp_c,
                streams[stream_id].target_temp_c,
                cold_in_c,
                cold_out_c,
            )
            > EPS
        )

        duty_kw = abs(h2 - h1)
        sum_hot_fcp = sum(abs(streams[stream_id].fcp_kw_per_k) for stream_id in hot_active)
        sum_cold_fcp = sum(abs(streams[stream_id].fcp_kw_per_k) for stream_id in cold_active)

        q_over_u = 0.0
        for stream_id in hot_active:
            stream = streams[stream_id]
            q_i = duty_kw * abs(stream.fcp_kw_per_k) / (sum_hot_fcp if sum_hot_fcp > EPS else 1.0)
            q_over_u += q_i / stream.overall_u_kw_m2_k
        for stream_id in cold_active:
            stream = streams[stream_id]
            q_i = duty_kw * abs(stream.fcp_kw_per_k) / (sum_cold_fcp if sum_cold_fcp > EPS else 1.0)
            q_over_u += q_i / stream.overall_u_kw_m2_k

        dt1 = hot_in_c - cold_out_c
        dt2 = hot_out_c - cold_in_c
        if dt1 <= EPS or dt2 <= EPS:
            lmtd = None
            area_m2 = None
        elif math.isclose(dt1, dt2, rel_tol=1e-9, abs_tol=1e-9):
            lmtd = (dt1 + dt2) / 2.0
            area_m2 = q_over_u / lmtd
        else:
            lmtd = abs((dt1 - dt2) / math.log(dt1 / dt2))
            area_m2 = q_over_u / lmtd

        intervals.append(
            AreaInterval(
                enthalpy_start_kw=h1,
                enthalpy_end_kw=h2,
                hot_in_c=hot_in_c,
                hot_out_c=hot_out_c,
                cold_in_c=cold_in_c,
                cold_out_c=cold_out_c,
                lmtd_c=lmtd,
                area_m2=area_m2,
                q_over_u_m2k=q_over_u,
                hot_stream_ids=hot_active,
                cold_stream_ids=cold_active,
            )
        )
    return tuple(intervals)


def curve_plot_records(
    streams: Sequence[ThermalStream],
    result: SupertargetingResult,
) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    hot_curve = _build_curve(
        streams,
        result.problem_table,
        curve_type="HOT",
        minimum_cold_utility_kw=result.minimum_cold_utility_kw,
    )
    hot_xs, hot_ys = _prepare_curve(hot_curve)

    for plot_name, cold_shift_kw in (
        ("composite_curve", 0.0),
        ("bcc_curve", result.minimum_cold_utility_kw),
    ):
        cold_curve = _build_curve(
            streams,
            result.problem_table,
            curve_type="COLD",
            minimum_cold_utility_kw=cold_shift_kw,
        )
        cold_xs, cold_ys = _prepare_curve(cold_curve)

        for point_index, (enthalpy_kw, temperature_c) in enumerate(zip(hot_xs, hot_ys), start=1):
            records.append(
                {
                    "plot_name": plot_name,
                    "curve_name": "hot_curve",
                    "point_index": point_index,
                    "enthalpy_kw": float(enthalpy_kw),
                    "temperature_c": float(temperature_c),
                    "cold_curve_shift_kw": float(cold_shift_kw),
                }
            )
        for point_index, (enthalpy_kw, temperature_c) in enumerate(zip(cold_xs, cold_ys), start=1):
            records.append(
                {
                    "plot_name": plot_name,
                    "curve_name": "cold_curve",
                    "point_index": point_index,
                    "enthalpy_kw": float(enthalpy_kw),
                    "temperature_c": float(temperature_c),
                    "cold_curve_shift_kw": float(cold_shift_kw),
                }
            )
    return records


def build_template_notebook_analysis(
    streams: Sequence[ThermalStream],
    *,
    delta_tmin_c: float,
    shift_cold_curve: bool = True,
) -> dict[str, object]:
    rows = _build_notebook_table(streams)
    hot_rows = [row for row in rows if row["Stream Type"] == "HOT"]
    cold_rows = [row for row in rows if row["Stream Type"] == "COLD"]

    temp_left = _unique_sorted_desc(
        [row["Ts"] for row in cold_rows]
        + [row["Tt"] for row in cold_rows]
        + [round(float(row["Ts"]) - delta_tmin_c, 2) for row in hot_rows]
        + [round(float(row["Tt"]) - delta_tmin_c, 2) for row in hot_rows]
    )
    temp_right = _unique_sorted_desc(
        [row["Ts"] for row in hot_rows]
        + [row["Tt"] for row in hot_rows]
        + [round(float(row["Ts"]) + delta_tmin_c, 2) for row in cold_rows]
        + [round(float(row["Tt"]) + delta_tmin_c, 2) for row in cold_rows]
    )
    if len(temp_left) != len(temp_right):
        raise RuntimeError("Template notebook calculation failed because temperature ladders differ in length.")

    stream_count = len(rows) + 1
    streams_involved = np.zeros((len(temp_left), stream_count), dtype=int)
    cold_streams_involved = np.zeros((len(temp_left), stream_count), dtype=int)
    hot_streams_involved = np.zeros((len(temp_left), stream_count), dtype=int)

    for i in range(len(temp_left)):
        for cold_row in cold_rows:
            active_id = np.where(cold_row["Ts"] <= temp_left[i] < cold_row["Tt"], cold_row["id"], 0)
            streams_involved[i][active_id] = active_id
            cold_streams_involved[i][active_id] = active_id
        for hot_row in hot_rows:
            active_id = np.where(hot_row["Tt"] <= temp_right[i] < hot_row["Ts"], hot_row["id"], 0)
            streams_involved[i][active_id] = active_id
            hot_streams_involved[i][active_id] = active_id

    differences = [0.0]
    for i in range(1, len(temp_left)):
        differences.append(temp_left[i - 1] - temp_left[i])

    sum_fcp: list[float] = []
    deficits: list[float] = []
    problem_table_records: list[dict[str, object]] = []
    for i in range(len(temp_left)):
        active_ids = [int(stream_id) for stream_id in streams_involved[i] if stream_id != 0]
        total = sum(float(rows[stream_id - 1]["FCp"]) for stream_id in active_ids)
        sum_fcp.append(total)
        deficit = total * differences[i]
        deficits.append(deficit)
        problem_table_records.append(
            {
                "subnetwork": f"SN{i}",
                "Tleft": float(temp_left[i]),
                "Tright": float(temp_right[i]),
                "streams_involved": tuple(active_ids),
                "temperature_difference_c": float(differences[i]),
                "sum_fcp_kw_per_k": float(total),
                "deficit_kw": float(deficit),
            }
        )

    accumulated_output: list[float] = []
    running = 0.0
    for deficit in deficits:
        running += -deficit
        accumulated_output.append(running)
    accumulated_input = [0.0] + [-accumulated_output[i - 1] for i in range(1, len(temp_left))]
    offset = abs(min(accumulated_output, default=0.0))
    heat_in = [offset + value for value in accumulated_input]
    heat_out = [offset + value for value in accumulated_output]

    for i in range(len(problem_table_records)):
        problem_table_records[i]["accumulated_input_kw"] = float(accumulated_input[i])
        problem_table_records[i]["accumulated_output_kw"] = float(accumulated_output[i])
        problem_table_records[i]["heat_flows_input_kw"] = float(heat_in[i])
        problem_table_records[i]["heat_flows_output_kw"] = float(heat_out[i])

    pinch_hot_candidates = [temp_right[i] for i in range(len(temp_right)) if math.isclose(heat_out[i], 0.0, abs_tol=1e-9)]
    if not pinch_hot_candidates:
        pinch_hot_candidates = [temp_right[min(range(len(heat_out)), key=lambda idx: abs(heat_out[idx]))]]
    pinch_hot_c = float(pinch_hot_candidates[0])
    minimum_hot_utility_kw = float(heat_in[0])
    minimum_cold_utility_kw = float(heat_out[-1])

    hot_curve_records: list[dict[str, object]] = []
    cumulative = 0.0
    reversed_differences = list(reversed(differences))
    reversed_tright = list(reversed(temp_right))
    for i in range(len(temp_left)):
        active_ids = tuple(int(stream_id) for stream_id in hot_streams_involved[len(temp_left) - i - 1] if stream_id != 0)
        total = sum(float(rows[stream_id - 1]["FCp"]) for stream_id in active_ids)
        hot_curve_records.append(
            {
                "Temperature": float(reversed_tright[i]),
                "streams_involved": active_ids,
                "sum_fcp_kw_per_k": float(total),
                "difference_c": float(reversed_differences[i]),
                "enthalpy_not_final_kw": float(reversed_differences[i] * total),
                "Final Enthalpy": float(abs(cumulative)),
            }
        )
        cumulative += reversed_differences[i] * total
    hot_valid_temps = {float(row["Ts"]) for row in hot_rows} | {float(row["Tt"]) for row in hot_rows}
    hot_curve_records = [row for row in hot_curve_records if row["Temperature"] in hot_valid_temps]

    cold_curve_records: list[dict[str, object]] = []
    cumulative = 0.0
    reversed_tleft = list(reversed(temp_left))
    for i in range(len(temp_left)):
        active_ids = tuple(int(stream_id) for stream_id in cold_streams_involved[len(temp_left) - i - 1] if stream_id != 0)
        total = sum(float(rows[stream_id - 1]["FCp"]) for stream_id in active_ids)
        cold_curve_records.append(
            {
                "Temperature": float(reversed_tleft[i]),
                "streams_involved": active_ids,
                "sum_fcp_kw_per_k": float(total),
                "difference_c": float(reversed_differences[i]),
                "enthalpy_not_final_kw": float(reversed_differences[i] * total),
                "Final Enthalpy": float(abs(cumulative) + (minimum_cold_utility_kw if shift_cold_curve else 0.0)),
            }
        )
        cumulative += reversed_differences[i] * total
    cold_valid_temps = {float(row["Ts"]) for row in cold_rows} | {float(row["Tt"]) for row in cold_rows}
    cold_curve_records = [row for row in cold_curve_records if row["Temperature"] in cold_valid_temps]

    input_table_records: list[dict[str, object]] = []
    for row in rows:
        input_table_records.append(
            {
                "Stream Number": int(row["id"]),
                "Stream Information": str(row["name"]),
                "Ts": float(row["Ts"]),
                "Tt": float(row["Tt"]),
                "Q": float(row["Q"]),
                "U": float(row["U"]),
                "FCp": float(row["FCp"]),
                "Stream Type": str(row["Stream Type"]),
            }
        )

    return {
        "delta_tmin_c": float(delta_tmin_c),
        "pinch_hot_c": pinch_hot_c,
        "minimum_hot_utility_kw": minimum_hot_utility_kw,
        "minimum_cold_utility_kw": minimum_cold_utility_kw,
        "input_table_records": input_table_records,
        "problem_table_records": problem_table_records,
        "hot_curve_records": hot_curve_records,
        "cold_curve_records": cold_curve_records,
    }


def save_template_combined_composite_curve_figure(
    including_utility_analysis: Mapping[str, object],
    excluding_utility_analysis: Mapping[str, object],
    output_path: str | Path,
) -> Path:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    output_path = Path(output_path)

    def _curve_frame(analysis: Mapping[str, object], key: str) -> list[dict[str, float]]:
        rows = analysis[key]
        assert isinstance(rows, list)
        return rows

    hot_curve_rows = _curve_frame(including_utility_analysis, "hot_curve_records")
    cold_curve_rows = _curve_frame(including_utility_analysis, "cold_curve_records")
    hot_curve_rows_process = _curve_frame(excluding_utility_analysis, "hot_curve_records")
    cold_curve_rows_process = _curve_frame(excluding_utility_analysis, "cold_curve_records")

    composite_fig, axes = plt.subplots(1, 2, figsize=(14, 6), constrained_layout=True)
    plot_sets = (
        (axes[0], hot_curve_rows, cold_curve_rows, "(a) Including Utility Streams"),
        (axes[1], hot_curve_rows_process, cold_curve_rows_process, "(b) Excluding Utility Streams"),
    )
    for ax, hot_rows_plot, cold_rows_plot, title in plot_sets:
        if not hot_rows_plot or not cold_rows_plot:
            continue
        ax.set_xlim(0, 1.1 * max(float(row["Final Enthalpy"]) for row in cold_rows_plot))
        ax.set_ylim(
            0,
            1.1
            * max(
                max(float(row["Temperature"]) for row in hot_rows_plot),
                max(float(row["Temperature"]) for row in cold_rows_plot),
            ),
        )
        ax.plot(
            [float(row["Final Enthalpy"]) for row in hot_rows_plot],
            [float(row["Temperature"]) for row in hot_rows_plot],
            color="red",
            label="Hot Composite Curve",
        )
        ax.plot(
            [float(row["Final Enthalpy"]) for row in cold_rows_plot],
            [float(row["Temperature"]) for row in cold_rows_plot],
            color="blue",
            label="Cold Composite Curve",
        )
        ax.legend()
        ax.set_xlabel("Enthalpy (kW)")
        ax.set_ylabel("Temperature (°C)")
        ax.set_title(title)

    composite_fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(composite_fig)
    return output_path


def replay_curve_plot_records(
    streams: Sequence[ThermalStream],
    *,
    delta_tmin_c: float,
) -> list[dict[str, object]]:
    rows = _build_notebook_table(streams)
    hot_rows = [row for row in rows if row["Stream Type"] == "HOT"]
    cold_rows = [row for row in rows if row["Stream Type"] == "COLD"]

    temp_left = _unique_sorted_desc(
        [row["Ts"] for row in cold_rows]
        + [row["Tt"] for row in cold_rows]
        + [round(float(row["Ts"]) - delta_tmin_c, 2) for row in hot_rows]
        + [round(float(row["Tt"]) - delta_tmin_c, 2) for row in hot_rows]
    )
    temp_right = _unique_sorted_desc(
        [row["Ts"] for row in hot_rows]
        + [row["Tt"] for row in hot_rows]
        + [round(float(row["Ts"]) + delta_tmin_c, 2) for row in cold_rows]
        + [round(float(row["Tt"]) + delta_tmin_c, 2) for row in cold_rows]
    )
    if len(temp_left) != len(temp_right):
        raise RuntimeError("Notebook replay failed because temperature ladders differ in length.")

    stream_count = len(rows) + 1
    cold_streams_involved = np.zeros((len(temp_left), stream_count), dtype=int)
    hot_streams_involved = np.zeros((len(temp_left), stream_count), dtype=int)

    for i in range(len(temp_left)):
        for cold_row in cold_rows:
            active_id = np.where(cold_row["Ts"] <= temp_left[i] < cold_row["Tt"], cold_row["id"], 0)
            cold_streams_involved[i][active_id] = active_id
        for hot_row in hot_rows:
            active_id = np.where(hot_row["Tt"] <= temp_right[i] < hot_row["Ts"], hot_row["id"], 0)
            hot_streams_involved[i][active_id] = active_id

    differences = [0.0]
    for i in range(1, len(temp_left)):
        differences.append(temp_left[i - 1] - temp_left[i])

    deficits: list[float] = []
    for i in range(len(temp_left)):
        total = 0.0
        for stream_id in cold_streams_involved[i]:
            if stream_id != 0:
                total += float(rows[stream_id - 1]["FCp"])
        for stream_id in hot_streams_involved[i]:
            if stream_id != 0:
                total += float(rows[stream_id - 1]["FCp"])
        deficits.append(total * differences[i])

    accumulated_output: list[float] = []
    running = 0.0
    for deficit in deficits:
        running += -deficit
        accumulated_output.append(running)
    minimum_cold_utility_kw = accumulated_output[-1] + abs(min(accumulated_output, default=0.0))

    records: list[dict[str, object]] = []

    cumulative = 0.0
    reversed_differences = list(reversed(differences))
    reversed_tright = list(reversed(temp_right))
    hot_valid_temps = {float(row["Ts"]) for row in hot_rows} | {float(row["Tt"]) for row in hot_rows}
    point_index = 1
    for i in range(len(temp_left)):
        temperature_c = float(reversed_tright[i])
        if temperature_c in hot_valid_temps:
            records.append(
                {
                    "plot_name": "bcc_curve",
                    "curve_name": "hot_curve",
                    "point_index": point_index,
                    "enthalpy_kw": float(abs(cumulative)),
                    "temperature_c": temperature_c,
                    "cold_curve_shift_kw": float(minimum_cold_utility_kw),
                }
            )
            point_index += 1
        active_ids = tuple(stream_id for stream_id in hot_streams_involved[len(temp_left) - i - 1] if stream_id != 0)
        total = sum(float(rows[stream_id - 1]["FCp"]) for stream_id in active_ids)
        cumulative += reversed_differences[i] * total

    cumulative = 0.0
    reversed_tleft = list(reversed(temp_left))
    cold_valid_temps = {float(row["Ts"]) for row in cold_rows} | {float(row["Tt"]) for row in cold_rows}
    point_index = 1
    for i in range(len(temp_left)):
        temperature_c = float(reversed_tleft[i])
        if temperature_c in cold_valid_temps:
            records.append(
                {
                    "plot_name": "bcc_curve",
                    "curve_name": "cold_curve",
                    "point_index": point_index,
                    "enthalpy_kw": float(abs(cumulative) + minimum_cold_utility_kw),
                    "temperature_c": temperature_c,
                    "cold_curve_shift_kw": float(minimum_cold_utility_kw),
                }
            )
            point_index += 1
        active_ids = tuple(stream_id for stream_id in cold_streams_involved[len(temp_left) - i - 1] if stream_id != 0)
        total = sum(float(rows[stream_id - 1]["FCp"]) for stream_id in active_ids)
        cumulative += reversed_differences[i] * total

    return records


def _count_streams_on_side(
    streams: Sequence[ThermalStream],
    *,
    hot_pinch_c: float,
    cold_pinch_c: float,
    side: str,
) -> tuple[int, int]:
    hot_count = 0
    cold_count = 0
    for stream in streams:
        low = min(stream.supply_temp_c, stream.target_temp_c)
        high = max(stream.supply_temp_c, stream.target_temp_c)
        if stream.stream_type == "HOT":
            if side == "ABOVE" and _overlap_length(low, high, hot_pinch_c, high) > EPS:
                hot_count += 1
            if side == "BELOW" and _overlap_length(low, high, low, hot_pinch_c) > EPS:
                hot_count += 1
        else:
            if side == "ABOVE" and _overlap_length(low, high, cold_pinch_c, high) > EPS:
                cold_count += 1
            if side == "BELOW" and _overlap_length(low, high, low, cold_pinch_c) > EPS:
                cold_count += 1
    return hot_count, cold_count


def calculate_minimum_exchangers(
    streams: Sequence[ThermalStream],
    *,
    hot_pinch_c: float,
    cold_pinch_c: float,
    minimum_hot_utility_kw: float,
    minimum_cold_utility_kw: float,
) -> tuple[int, int, int]:
    hot_above, cold_above = _count_streams_on_side(
        streams,
        hot_pinch_c=hot_pinch_c,
        cold_pinch_c=cold_pinch_c,
        side="ABOVE",
    )
    hot_below, cold_below = _count_streams_on_side(
        streams,
        hot_pinch_c=hot_pinch_c,
        cold_pinch_c=cold_pinch_c,
        side="BELOW",
    )
    units_above = max(
        hot_above + cold_above + (1 if minimum_hot_utility_kw > EPS else 0) - 1,
        0,
    )
    units_below = max(
        hot_below + cold_below + (1 if minimum_cold_utility_kw > EPS else 0) - 1,
        0,
    )
    return units_above + units_below, units_above, units_below


def run_supertargeting(
    streams: Sequence[ThermalStream],
    *,
    delta_tmin_c: float,
) -> SupertargetingResult:
    if delta_tmin_c <= 0.0:
        raise ValueError("delta_tmin_c must be positive.")

    problem_rows = build_problem_table(streams, delta_tmin_c)
    minimum_hot_utility_kw = problem_rows[0].heat_in_kw
    minimum_cold_utility_kw = problem_rows[-1].heat_out_kw
    pinch_hot_c, pinch_cold_c = _find_pinch(problem_rows, delta_tmin_c)
    minimum_exchangers, above, below = calculate_minimum_exchangers(
        streams,
        hot_pinch_c=pinch_hot_c,
        cold_pinch_c=pinch_cold_c,
        minimum_hot_utility_kw=minimum_hot_utility_kw,
        minimum_cold_utility_kw=minimum_cold_utility_kw,
    )
    area_intervals = build_area_intervals(streams, problem_rows, minimum_cold_utility_kw)
    minimum_area_m2 = sum(interval.area_m2 or 0.0 for interval in area_intervals)
    return SupertargetingResult(
        delta_tmin_c=delta_tmin_c,
        pinch_hot_c=pinch_hot_c,
        pinch_cold_c=pinch_cold_c,
        minimum_hot_utility_kw=minimum_hot_utility_kw,
        minimum_cold_utility_kw=minimum_cold_utility_kw,
        minimum_exchangers=minimum_exchangers,
        minimum_exchangers_above_pinch=above,
        minimum_exchangers_below_pinch=below,
        minimum_area_m2=minimum_area_m2,
        problem_table=tuple(problem_rows),
        area_intervals=tuple(area_intervals),
    )


def _interp_temperature(enthalpy: float, xs: Sequence[float], ys: Sequence[float]) -> float:
    if not xs:
        raise ValueError("Interpolation curve is empty.")
    return float(np.interp(enthalpy, xs, ys))


def _interp_temperature_nan(enthalpy: float, xs: Sequence[float], ys: Sequence[float]) -> float:
    if not xs:
        return float("nan")
    lower = min(xs)
    upper = max(xs)
    if enthalpy < lower - EPS or enthalpy > upper + EPS:
        return float("nan")
    return float(np.interp(enthalpy, xs, ys))


def replay_notebook_area_algorithm(
    streams: Sequence[ThermalStream],
    *,
    delta_tmin_c: float,
) -> NotebookReplayResult:
    rows = _build_notebook_table(streams)
    hot_rows = [row for row in rows if row["Stream Type"] == "HOT"]
    cold_rows = [row for row in rows if row["Stream Type"] == "COLD"]

    temp_left = _unique_sorted_desc(
        [row["Ts"] for row in cold_rows]
        + [row["Tt"] for row in cold_rows]
        + [round(float(row["Ts"]) - delta_tmin_c, 2) for row in hot_rows]
        + [round(float(row["Tt"]) - delta_tmin_c, 2) for row in hot_rows]
    )
    temp_right = _unique_sorted_desc(
        [row["Ts"] for row in hot_rows]
        + [row["Tt"] for row in hot_rows]
        + [round(float(row["Ts"]) + delta_tmin_c, 2) for row in cold_rows]
        + [round(float(row["Tt"]) + delta_tmin_c, 2) for row in cold_rows]
    )
    if len(temp_left) != len(temp_right):
        raise RuntimeError("Notebook replay failed because temperature ladders differ in length.")

    stream_count = len(rows) + 1
    streams_involved = np.zeros((len(temp_left), stream_count), dtype=int)
    cold_streams_involved = np.zeros((len(temp_left), stream_count), dtype=int)
    hot_streams_involved = np.zeros((len(temp_left), stream_count), dtype=int)

    for i in range(len(temp_left)):
        for cold_row in cold_rows:
            active_id = np.where(cold_row["Ts"] <= temp_left[i] < cold_row["Tt"], cold_row["id"], 0)
            streams_involved[i][active_id] = active_id
            cold_streams_involved[i][active_id] = active_id
        for hot_row in hot_rows:
            active_id = np.where(hot_row["Tt"] <= temp_right[i] < hot_row["Ts"], hot_row["id"], 0)
            streams_involved[i][active_id] = active_id
            hot_streams_involved[i][active_id] = active_id

    differences = [0.0]
    for i in range(1, len(temp_left)):
        differences.append(temp_left[i - 1] - temp_left[i])

    sum_fcp: list[float] = []
    deficits: list[float] = []
    for i in range(len(temp_left)):
        active_ids = [stream_id for stream_id in streams_involved[i] if stream_id != 0]
        total = sum(float(rows[stream_id - 1]["FCp"]) for stream_id in active_ids)
        sum_fcp.append(total)
        deficits.append(total * differences[i])

    accumulated_output: list[float] = []
    running = 0.0
    for deficit in deficits:
        running += -deficit
        accumulated_output.append(running)
    accumulated_input = [0.0] + [-accumulated_output[i - 1] for i in range(1, len(temp_left))]
    offset = abs(min(accumulated_output, default=0.0))
    heat_in = [offset + value for value in accumulated_input]
    heat_out = [offset + value for value in accumulated_output]

    pinch_hot_candidates = [temp_right[i] for i in range(len(temp_right)) if math.isclose(heat_out[i], 0.0, abs_tol=1e-9)]
    if not pinch_hot_candidates:
        pinch_hot_candidates = [temp_right[min(range(len(heat_out)), key=lambda idx: abs(heat_out[idx]))]]
    pinch_hot_c = pinch_hot_candidates[0]
    minimum_hot_utility_kw = heat_in[0]
    minimum_cold_utility_kw = heat_out[-1]

    hot_curve_rows: list[dict[str, object]] = []
    cumulative = 0.0
    reversed_differences = list(reversed(differences))
    reversed_tright = list(reversed(temp_right))
    for i in range(len(temp_left)):
        active_ids = tuple(stream_id for stream_id in hot_streams_involved[len(temp_left) - i - 1] if stream_id != 0)
        total = sum(float(rows[stream_id - 1]["FCp"]) for stream_id in active_ids)
        hot_curve_rows.append(
            {
                "Temperature": np.float32(reversed_tright[i]),
                "Streams Involved": active_ids,
                "Final Enthalpy": np.float32(abs(cumulative)),
            }
        )
        cumulative += reversed_differences[i] * total
    hot_valid_temps = {float(row["Ts"]) for row in hot_rows} | {float(row["Tt"]) for row in hot_rows}
    hot_curve_rows = [row for row in hot_curve_rows if row["Temperature"] in hot_valid_temps]

    cold_curve_rows: list[dict[str, object]] = []
    cumulative = 0.0
    reversed_tleft = list(reversed(temp_left))
    for i in range(len(temp_left)):
        active_ids = tuple(stream_id for stream_id in cold_streams_involved[len(temp_left) - i - 1] if stream_id != 0)
        total = sum(float(rows[stream_id - 1]["FCp"]) for stream_id in active_ids)
        cold_curve_rows.append(
            {
                "Temperature": np.float32(reversed_tleft[i]),
                "Streams Involved": active_ids,
                "Final Enthalpy": np.float32(abs(cumulative) + minimum_cold_utility_kw),
            }
        )
        cumulative += reversed_differences[i] * total
    cold_valid_temps = {float(row["Ts"]) for row in cold_rows} | {float(row["Tt"]) for row in cold_rows}
    cold_curve_rows = [row for row in cold_curve_rows if row["Temperature"] in cold_valid_temps]

    hot_x = [float(row["Final Enthalpy"]) for row in hot_curve_rows]
    hot_y = [float(row["Temperature"]) for row in hot_curve_rows]
    cold_x = [float(row["Final Enthalpy"]) for row in cold_curve_rows]
    cold_y = [float(row["Temperature"]) for row in cold_curve_rows]

    hot_duplicate = [{"Temperature of Hot Streams": float(row["Temperature"]), "Final Enthalpy": float(row["Final Enthalpy"])} for row in hot_curve_rows]
    for enthalpy in cold_x:
        hot_duplicate.append(
            {
                "Temperature of Hot Streams": _interp_temperature_nan(enthalpy, hot_x, hot_y),
                "Final Enthalpy": float(enthalpy),
            }
        )

    cold_duplicate = [{"Temperature of Cold Streams": float(row["Temperature"]), "Final Enthalpy": float(row["Final Enthalpy"])} for row in cold_curve_rows]
    for enthalpy in hot_x:
        cold_duplicate.append(
            {
                "Temperature of Cold Streams": _interp_temperature_nan(enthalpy, cold_x, cold_y),
                "Final Enthalpy": float(enthalpy),
            }
        )

    def _drop_duplicates_keep_first(
        rows_in: Sequence[dict[str, float]],
        *,
        value_key: str,
        temp_key: str,
    ) -> list[dict[str, float]]:
        unique_rows: list[dict[str, float]] = []
        seen: list[tuple[float, float]] = []
        for row in rows_in:
            key = (float(row[temp_key]), float(row[value_key]))
            is_duplicate = False
            for existing in seen:
                if (
                    math.isclose(existing[0], key[0], abs_tol=EPS) or (math.isnan(existing[0]) and math.isnan(key[0]))
                ) and (
                    math.isclose(existing[1], key[1], abs_tol=EPS) or (math.isnan(existing[1]) and math.isnan(key[1]))
                ):
                    is_duplicate = True
                    break
            if not is_duplicate:
                seen.append(key)
                unique_rows.append(row)
        return unique_rows

    hot_duplicate_sorted = sorted(
        _drop_duplicates_keep_first(
            hot_duplicate,
            value_key="Final Enthalpy",
            temp_key="Temperature of Hot Streams",
        ),
        key=lambda row: row["Final Enthalpy"],
    )
    cold_duplicate_sorted = sorted(
        _drop_duplicates_keep_first(
            cold_duplicate,
            value_key="Final Enthalpy",
            temp_key="Temperature of Cold Streams",
        ),
        key=lambda row: row["Final Enthalpy"],
    )

    minimum_area_rows: list[dict[str, float]] = []
    row_count = max(len(hot_duplicate_sorted), len(cold_duplicate_sorted))
    combined_rows: list[dict[str, float]] = []
    for i in range(row_count):
        hot_temp = float(hot_duplicate_sorted[i]["Temperature of Hot Streams"]) if i < len(hot_duplicate_sorted) else float("nan")
        hot_enthalpy = float(hot_duplicate_sorted[i]["Final Enthalpy"]) if i < len(hot_duplicate_sorted) else float("nan")
        cold_temp = float(cold_duplicate_sorted[i]["Temperature of Cold Streams"]) if i < len(cold_duplicate_sorted) else float("nan")
        cold_enthalpy = float(cold_duplicate_sorted[i]["Final Enthalpy"]) if i < len(cold_duplicate_sorted) else float("nan")
        combined_rows.append(
            {
                "Temperature of Hot Streams": hot_temp,
                "Temperature of Cold Streams": cold_temp,
                "Final Enthalpy of either streams": cold_enthalpy,
                "_hot_enthalpy": hot_enthalpy,
            }
        )

    seen_enthalpy: set[float] = set()
    for row in combined_rows:
        enthalpy = row["Final Enthalpy of either streams"]
        hot_temp = row["Temperature of Hot Streams"]
        cold_temp = row["Temperature of Cold Streams"]
        hot_enthalpy = row["_hot_enthalpy"]
        if any(math.isnan(value) for value in (enthalpy, hot_temp, cold_temp, hot_enthalpy)):
            continue
        rounded_enthalpy = round(enthalpy, 10)
        if rounded_enthalpy in seen_enthalpy:
            continue
        seen_enthalpy.add(rounded_enthalpy)
        minimum_area_rows.append(
            {
                "Temperature of Hot Streams": hot_temp,
                "Temperature of Cold Streams": cold_temp,
                "Final Enthalpy of either streams": enthalpy,
            }
        )

    auto_hot_ids: dict[int, list[int]] = {}
    auto_cold_ids: dict[int, list[int]] = {}
    for n in range(len(minimum_area_rows) - 1):
        th1 = float(minimum_area_rows[n]["Temperature of Hot Streams"])
        tc1 = float(minimum_area_rows[n]["Temperature of Cold Streams"])
        th2 = float(minimum_area_rows[n + 1]["Temperature of Hot Streams"])
        tc2 = float(minimum_area_rows[n + 1]["Temperature of Cold Streams"])

        hot_ids = [
            int(row["id"])
            for row in hot_rows
            if float(row["Tt"]) < max(th1, th2) and float(row["Ts"]) > min(th1, th2)
        ]
        cold_ids = [
            int(row["id"])
            for row in cold_rows
            if float(row["Ts"]) < max(tc1, tc2) and float(row["Tt"]) > min(tc1, tc2)
        ]
        auto_hot_ids[n] = hot_ids
        auto_cold_ids[n] = cold_ids

    segments: list[NotebookAreaSegment] = []
    for n in range(len(minimum_area_rows) - 1):
        th1 = float(minimum_area_rows[n]["Temperature of Hot Streams"])
        tc1 = float(minimum_area_rows[n]["Temperature of Cold Streams"])
        th2 = float(minimum_area_rows[n + 1]["Temperature of Hot Streams"])
        tc2 = float(minimum_area_rows[n + 1]["Temperature of Cold Streams"])
        hot_in = max(th1, th2)
        hot_out = min(th1, th2)
        cold_in = min(tc1, tc2)
        cold_out = max(tc1, tc2)

        hot_ids = auto_hot_ids[n]
        cold_ids = auto_cold_ids[n]

        for stream_id in hot_ids:
            row = rows[stream_id - 1]
            ts = float(row["Ts"])
            tt = float(row["Tt"])
            if min(th1, th2) <= tt <= max(th1, th2) and min(th1, th2) <= ts <= max(th1, th2):
                hot_in = ts
                hot_out = tt

        delta_tb = th1 - tc1
        delta_ta = th2 - tc2
        if math.isclose(delta_ta, delta_tb, abs_tol=1e-12):
            lmtd = delta_ta
        else:
            a = math.copysign(max(1e-9, abs(delta_ta)), delta_ta)
            b = math.copysign(max(1e-9, abs(delta_tb)), delta_tb)
            lmtd = abs((a - b) / math.log(a / b))

        dth = th1 - th2
        dtc = tc2 - tc1
        shot = 0.0
        scold = 0.0

        for stream_id in hot_ids:
            row = rows[stream_id - 1]
            u_value = float(row["U"])
            if u_value <= 0.0 or math.isnan(u_value):
                continue
            ts = float(row["Ts"])
            tt = float(row["Tt"])
            if math.isclose(hot_in, ts, abs_tol=1e-6) and math.isclose(hot_out, tt, abs_tol=1e-6):
                dt_local = ts - tt
            else:
                dt_local = abs(dth)
            q_i = float(row["FCp"]) * dt_local
            shot += abs(q_i / u_value)

        for stream_id in cold_ids:
            row = rows[stream_id - 1]
            u_value = float(row["U"])
            if u_value <= 0.0 or math.isnan(u_value):
                continue
            q_j = float(row["FCp"]) * abs(dtc)
            scold += abs(q_j / u_value)

        area_m2 = 0.0 if lmtd in (None, 0.0) or math.isnan(lmtd) else (abs(shot) + abs(scold)) / lmtd
        hot_streams = tuple(
            (stream_id, abs(float(rows[stream_id - 1]["FCp"])), float(rows[stream_id - 1]["U"]))
            for stream_id in hot_ids
        )
        cold_streams = tuple(
            (stream_id, abs(float(rows[stream_id - 1]["FCp"])), float(rows[stream_id - 1]["U"]))
            for stream_id in cold_ids
        )
        segments.append(
            NotebookAreaSegment(
                segment=n,
                hot_in_c=hot_in,
                cold_in_c=cold_in,
                hot_out_c=hot_out,
                cold_out_c=cold_out,
                delta_ta_c=delta_ta,
                delta_tb_c=delta_tb,
                lmtd_c=lmtd,
                hot_q_over_u_m2k=abs(shot),
                cold_q_over_u_m2k=abs(scold),
                area_m2=area_m2,
                hot_streams=hot_streams,
                cold_streams=cold_streams,
            )
        )

    total_area_m2 = float(sum(segment.area_m2 for segment in segments))
    return NotebookReplayResult(
        delta_tmin_c=delta_tmin_c,
        pinch_hot_c=float(pinch_hot_c),
        minimum_hot_utility_kw=float(minimum_hot_utility_kw),
        minimum_cold_utility_kw=float(minimum_cold_utility_kw),
        total_area_m2=total_area_m2,
        segments=tuple(segments),
    )


def sweep_delta_tmin(
    streams: Sequence[ThermalStream],
    delta_tmin_values_c: Sequence[float],
) -> list[SupertargetingResult]:
    return [run_supertargeting(streams, delta_tmin_c=value) for value in delta_tmin_values_c]


def _format_stream_ids(streams: Sequence[ThermalStream], stream_ids: Sequence[int]) -> str:
    return ", ".join(streams[stream_id].name for stream_id in stream_ids) if stream_ids else "-"


def print_stream_table(streams: Sequence[ThermalStream]) -> None:
    print("Loaded thermal streams")
    print("-" * 112)
    print(
        f"{'Name':40s} {'Type':6s} {'Ts (C)':>10s} {'Tt (C)':>10s} "
        f"{'Q (kW)':>12s} {'U (kW/m2.K)':>14s} {'FCp (kW/K)':>12s}"
    )
    print("-" * 112)
    for stream in streams:
        print(
            f"{stream.name[:40]:40s} {stream.stream_type:6s} "
            f"{stream.supply_temp_c:10.3f} {stream.target_temp_c:10.3f} "
            f"{stream.heat_load_kw:12.3f} {stream.overall_u_kw_m2_k:14.3f} "
            f"{stream.fcp_kw_per_k:12.3f}"
        )
    print()


def print_summary(result: SupertargetingResult) -> None:
    print(f"Supertargeting summary at ΔTmin = {result.delta_tmin_c:.3f} C")
    print("-" * 72)
    print(f"Hot pinch temperature (C):      {result.pinch_hot_c:12.3f}")
    print(f"Cold pinch temperature (C):     {result.pinch_cold_c:12.3f}")
    print(f"Minimum hot utility (kW):       {result.minimum_hot_utility_kw:12.3f}")
    print(f"Minimum cold utility (kW):      {result.minimum_cold_utility_kw:12.3f}")
    print(f"Minimum exchangers above pinch: {result.minimum_exchangers_above_pinch:12d}")
    print(f"Minimum exchangers below pinch: {result.minimum_exchangers_below_pinch:12d}")
    print(f"Minimum exchangers total:       {result.minimum_exchangers:12d}")
    print(f"Minimum total area (m2):        {result.minimum_area_m2:12.3f}")
    print()


def print_area_table(result: SupertargetingResult, streams: Sequence[ThermalStream]) -> None:
    print("Area intervals")
    print("-" * 168)
    print(
        f"{'ΔH start':>10s} {'ΔH end':>10s} {'Hot in':>10s} {'Hot out':>10s} "
        f"{'Cold in':>10s} {'Cold out':>10s} {'LMTD':>10s} {'Area':>10s} "
        f"{'Hot streams':30s} {'Cold streams':30s}"
    )
    print("-" * 168)
    for interval in result.area_intervals:
        lmtd_text = f"{interval.lmtd_c:.3f}" if interval.lmtd_c is not None else "NA"
        area_text = f"{interval.area_m2:.3f}" if interval.area_m2 is not None else "NA"
        print(
            f"{interval.enthalpy_start_kw:10.3f} {interval.enthalpy_end_kw:10.3f} "
            f"{interval.hot_in_c:10.3f} {interval.hot_out_c:10.3f} "
            f"{interval.cold_in_c:10.3f} {interval.cold_out_c:10.3f} "
            f"{lmtd_text:>10s} {area_text:>10s} "
            f"{_format_stream_ids(streams, interval.hot_stream_ids)[:30]:30s} "
            f"{_format_stream_ids(streams, interval.cold_stream_ids)[:30]:30s}"
        )
    print()


def print_notebook_replay(result: NotebookReplayResult) -> None:
    print(f"Notebook replay summary at ΔTmin = {result.delta_tmin_c:.3f} C")
    print("-" * 72)
    print(f"Pinch hot temperature (C):      {result.pinch_hot_c:12.3f}")
    print(f"Minimum hot utility (kW):       {result.minimum_hot_utility_kw:12.3f}")
    print(f"Minimum cold utility (kW):      {result.minimum_cold_utility_kw:12.3f}")
    print(f"Notebook total area (m2):       {result.total_area_m2:12.3f}")
    print()
    print("Notebook replay area segments")
    print("-" * 156)
    print(
        f"{'Seg':>4s} {'Th_in':>12s} {'Tc_in':>12s} {'Th_out':>12s} {'Tc_out':>12s} "
        f"{'ΔTa':>12s} {'ΔTb':>12s} {'LMTD':>12s} {'Σqhot/U':>12s} {'Σqcold/U':>12s} {'ΔA':>12s}"
    )
    print("-" * 156)
    for segment in result.segments:
        print(
            f"{segment.segment:4d} "
            f"{segment.hot_in_c:12.6f} {segment.cold_in_c:12.6f} "
            f"{segment.hot_out_c:12.6f} {segment.cold_out_c:12.6f} "
            f"{segment.delta_ta_c:12.6f} {segment.delta_tb_c:12.6f} "
            f"{segment.lmtd_c:12.6f} {segment.hot_q_over_u_m2k:12.6f} "
            f"{segment.cold_q_over_u_m2k:12.6f} {segment.area_m2:12.6f}"
        )
        print(f"     hot  = {list(segment.hot_streams)}")
        print(f"     cold = {list(segment.cold_streams)}")
    print()


def notebook_replay_summary_record(result: NotebookReplayResult) -> dict[str, float]:
    return {
        "delta_tmin_c": result.delta_tmin_c,
        "pinch_hot_c": result.pinch_hot_c,
        "minimum_hot_utility_kw": result.minimum_hot_utility_kw,
        "minimum_cold_utility_kw": result.minimum_cold_utility_kw,
        "total_area_m2": result.total_area_m2,
        "segment_count": float(len(result.segments)),
    }


def notebook_replay_segment_records(result: NotebookReplayResult) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for segment in result.segments:
        records.append(
            {
                "segment": segment.segment,
                "Th_in": segment.hot_in_c,
                "Tc_in": segment.cold_in_c,
                "Th_out": segment.hot_out_c,
                "Tc_out": segment.cold_out_c,
                "delta_ta": segment.delta_ta_c,
                "delta_tb": segment.delta_tb_c,
                "lmtd": segment.lmtd_c,
                "hot_q_over_u": segment.hot_q_over_u_m2k,
                "cold_q_over_u": segment.cold_q_over_u_m2k,
                "area_m2": segment.area_m2,
                "hot_streams": list(segment.hot_streams),
                "cold_streams": list(segment.cold_streams),
            }
        )
    return records


def supertargeting_summary_record(result: SupertargetingResult) -> dict[str, float]:
    return {
        "delta_tmin_c": result.delta_tmin_c,
        "pinch_hot_c": result.pinch_hot_c,
        "pinch_cold_c": result.pinch_cold_c,
        "minimum_hot_utility_kw": result.minimum_hot_utility_kw,
        "minimum_cold_utility_kw": result.minimum_cold_utility_kw,
        "minimum_exchangers": float(result.minimum_exchangers),
        "minimum_exchangers_above_pinch": float(result.minimum_exchangers_above_pinch),
        "minimum_exchangers_below_pinch": float(result.minimum_exchangers_below_pinch),
        "minimum_area_m2": result.minimum_area_m2,
    }


def _load_streams_from_args(args: argparse.Namespace) -> list[ThermalStream]:
    if args.input_xlsx:
        return read_streams_from_xlsx(args.input_xlsx, sheet_name=args.sheet)
    if args.input_csv:
        return read_streams_from_csv(args.input_csv)
    return read_streams_from_hysys(DEFAULT_CFP04_HYSYS_SPECS)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run pinch/supertargeting calculations.")
    parser.add_argument(
        "--delta-tmin",
        nargs="+",
        type=float,
        default=[5.0],
        help="One or more ΔTmin values in degree C.",
    )
    parser.add_argument(
        "--input-xlsx",
        type=Path,
        help="Read the thermal stream table from an .xlsx file instead of HYSYS.",
    )
    parser.add_argument(
        "--input-csv",
        type=Path,
        help="Read the thermal stream table from a .csv file instead of HYSYS.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Optional Excel sheet name when --input-xlsx is used.",
    )
    parser.add_argument(
        "--print-area-table",
        action="store_true",
        help="Print the detailed interval-by-interval area table.",
    )
    parser.add_argument(
        "--replay-notebook-area",
        action="store_true",
        help="Replay the original notebook area algorithm as closely as possible.",
    )
    parser.add_argument(
        "--exclude-standalone-utilities",
        action="store_true",
        help="Drop standalone utility rows such as MP/HP/CW/ChW/Refrigeration/LPSG before calculation.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    streams = _load_streams_from_args(args)
    if args.exclude_standalone_utilities:
        streams = filter_out_standalone_utility_streams(streams)
    print_stream_table(streams)
    if args.replay_notebook_area:
        for delta_tmin in args.delta_tmin:
            replay = replay_notebook_area_algorithm(streams, delta_tmin_c=delta_tmin)
            print_notebook_replay(replay)
        return 0
    results = sweep_delta_tmin(streams, args.delta_tmin)
    for result in results:
        print_summary(result)
        if args.print_area_table:
            print_area_table(result, streams)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
